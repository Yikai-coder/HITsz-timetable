# coding: utf-8

import config
import openpyxl
import re
import misc

from classParser import lexer, syntaxParser


def process():
    workbook = openpyxl.load_workbook(config.excel_file_path)
    worksheet= workbook.active

    orig_col = 0
    orig_row = 0

    try:
        for i in range(1, 10):
            for col in range(1, i+1):
                if worksheet.cell(col, i).value == "星期一":
                    orig_col, orig_row = col, i-1
                    raise misc.StopLoop
            for row in range(i, 0, -1):
                if worksheet.cell(i, row).value == "星期一":
                    orig_col, orig_row = i, row-1
                    raise misc.StopLoop
    except misc.StopLoop:
        pass


    # cal_data (calendar data): array of
    #   - Days (0=>Mon., 1=>Tue., ...): array of
    #       - time (0=>8:30-10:15, 1=>10:30-12:15, ...): array of
    #           - classes (class information): dict containing
    #               - classname
    #               - classroom
    #               - valid week
    cal_data = []

    for i in range(1, 8):
        every_day = []
        
        for j in range(1, 7):
            lexer_lst = lexer.parse(worksheet.cell(orig_col+j, orig_row+i).value)
            every_day.append(syntaxParser.parse(lexer_lst))

        cal_data.append(every_day)
    
    return cal_data